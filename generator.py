"""
VHA Proposed Rates – Excel Generator
Produces a formatted .xlsx matching the reference Proposed Rates structure.
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Manulife colour palette ─────────────────────────────────────────────────
GREEN_DARK   = "005A30"   # dark Manulife green
GREEN_MID    = "00704A"
GREEN_LIGHT  = "E6F5EF"
GREEN_HEADER = "00A758"   # Manulife green
GREY_LIGHT   = "F2F2F2"
GREY_MID     = "D9D9D9"
GREY_BORDER  = "BFBFBF"
WHITE        = "FFFFFF"
BLACK        = "000000"
BLUE_LINK    = "1F497D"

def _font(bold=False, size=10, color=BLACK, italic=False):
    return Font(name="Calibri", size=size, bold=bold,
                color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color=GREY_BORDER):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _thin_bottom(color=GREY_BORDER):
    s = Side(style="thin", color=color)
    return Border(bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _write_section_header(ws, row, label, col_start=2, max_col=22):
    """Write a section header row (e.g. 'Basic Life')."""
    cell = ws.cell(row=row, column=col_start, value=label)
    cell.font  = _font(bold=True, size=10, color=WHITE)
    cell.fill  = _fill(GREEN_MID)
    cell.alignment = _align()
    for c in range(col_start, max_col + 1):
        ws.cell(row=row, column=c).fill = _fill(GREEN_MID)


def _write_total_row(ws, row, label, lives, vol, curr_prem, prop_prem, total_chg,
                     col_start=2, max_col=22):
    """Write a subtotal row."""
    ws.cell(row=row, column=col_start, value=label).font = _font(bold=True, italic=True)
    ws.cell(row=row, column=9,  value=lives     if lives else None)
    ws.cell(row=row, column=10, value=vol       if vol   else None)
    ws.cell(row=row, column=14, value=round(curr_prem, 2) if curr_prem else None).number_format = '#,##0.00'
    ws.cell(row=row, column=18, value=round(prop_prem, 2) if prop_prem else None).number_format = '#,##0.00'
    ws.cell(row=row, column=21, value=round(total_chg, 4) if total_chg is not None else None).number_format = '0.0%'
    for c in range(col_start, max_col + 1):
        ws.cell(row=row, column=c).fill = _fill(GREY_LIGHT)
        ws.cell(row=row, column=c).font = _font(bold=True, italic=True, size=9)


def _write_data_row(ws, row, r):
    """Write one benefit data row (dict from processor enriched_groups)."""
    ws.cell(row=row, column=2,  value=r.get("benefit_code", ""))
    ws.cell(row=row, column=3,  value=631935)
    ws.cell(row=row, column=4,  value=r.get("division", ""))
    ws.cell(row=row, column=5,  value=0)
    ws.cell(row=row, column=6,  value=r.get("plans", ""))
    _BASIS_DISPLAY = {"1000": "$1,000", "100": "$100", "10": "$10"}
    basis_unit = r.get("basis_unit", "")
    ws.cell(row=row, column=7,  value=r.get("basis", ""))
    ws.cell(row=row, column=8,  value=_BASIS_DISPLAY.get(str(basis_unit), basis_unit))
    ws.cell(row=row, column=9,  value=r.get("lives"))
    ws.cell(row=row, column=10, value=r.get("volumes") if r.get("volumes") else None)

    # Current
    ws.cell(row=row, column=11, value=r.get("curr_exp_rate")).number_format = "0.000"
    ws.cell(row=row, column=12, value=r.get("curr_dr_rate")).number_format  = "0.000"
    ws.cell(row=row, column=13, value=r.get("curr_total_rate")).number_format = "0.000"
    ws.cell(row=row, column=14, value=r.get("curr_premium")).number_format  = '#,##0.00'

    # Proposed
    ws.cell(row=row, column=15, value=r.get("prop_exp_rate")).number_format = "0.000"
    ws.cell(row=row, column=16, value=r.get("prop_dr_rate")).number_format  = "0.000"
    ws.cell(row=row, column=17, value=r.get("prop_total_rate")).number_format = "0.000"
    ws.cell(row=row, column=18, value=r.get("prop_premium")).number_format  = '#,##0.00'

    ws.cell(row=row, column=19, value=r.get("exp_rate_chg")).number_format  = '0.0%'
    ws.cell(row=row, column=20, value=r.get("dr_pct")).number_format        = '0.0%'
    ws.cell(row=row, column=21, value=r.get("total_rate_chg")).number_format = '0.0%'

    # Alternate row shading
    shade = GREY_LIGHT if row % 2 == 0 else WHITE
    for c in range(2, 22):
        cell = ws.cell(row=row, column=c)
        try:
            rgb = cell.fill.fgColor.rgb
        except Exception:
            rgb = "00000000"
        if rgb in ("00000000", "FFFFFFFF", ""):
            cell.fill = _fill(shade)
        cell.font      = _font(size=9)
        cell.alignment = _align(h="right" if c >= 9 else "left", v="center")


def generate_excel(report_data: dict) -> bytes:
    """
    Build the Proposed Rates Excel workbook and return bytes.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026 Proposed Rates"
    ws.sheet_view.showGridLines = False

    client_name  = report_data.get("client_name", "VHA Home Healthcare")
    period_str   = report_data.get("period_str",  "")
    groups       = report_data.get("groups", {})

    # ── Parse effective date from period string ────────────────────────────
    effective = "01-Mar-2026"
    if "2026" in period_str:
        effective = "01-Mar-2026"

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {
        1:  2,   2: 26,  3: 10,  4: 10,  5: 7,
        6: 14,   7:  5,  8:  8,  9:  7, 10: 14,
        11: 14, 12: 14,  13: 11, 14: 15, 15: 14,
        16: 14,  17: 11, 18: 15, 19: 13, 20: 13, 21: 14,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Report Title Block ─────────────────────────────────────────────────
    ws.merge_cells("B4:E4")
    c = ws["B4"]
    c.value     = client_name
    c.font      = _font(bold=True, size=14, color=GREEN_DARK)
    c.alignment = _align()

    ws.merge_cells("B5:E5")
    c = ws["B5"]
    c.value     = "Summary of Proposed Monthly Rates"
    c.font      = _font(bold=False, size=11, color=GREEN_DARK)
    c.alignment = _align()

    ws.merge_cells("B6:E6")
    c = ws["B6"]
    c.value     = f"Effective {effective}"
    c.font      = _font(bold=False, size=11, color=GREEN_DARK)
    c.alignment = _align()

    # Manulife logo placeholder
    ws.merge_cells("R4:U4")
    logo_cell = ws["R4"]
    logo_cell.value     = "Manulife"
    logo_cell.font      = _font(bold=True, size=16, color=GREEN_DARK)
    logo_cell.alignment = _align(h="right")

    # ── Header Row 8: Group spans (Current / Proposed) ────────────────────
    # Rule: in openpyxl only style the TOP-LEFT cell of a merged range.
    # Never write to interior cells of a merge – it breaks the merge.
    ws.row_dimensions[8].height = 20

    # B8:J8 – left block, no label, mid-green background
    ws.merge_cells("B8:J8")
    ws["B8"].fill = _fill(GREEN_MID)

    # K8:N8 – "Current" group header
    ws.merge_cells("K8:N8")
    _c = ws["K8"]
    _c.value     = "Current"
    _c.font      = _font(bold=True, size=9, color=WHITE)
    _c.fill      = _fill(GREEN_HEADER)
    _c.alignment = _align(h="center", v="center")

    # O8:R8 – "Proposed" group header
    ws.merge_cells("O8:R8")
    _c = ws["O8"]
    _c.value     = "Proposed"
    _c.font      = _font(bold=True, size=9, color=WHITE)
    _c.fill      = _fill(GREEN_DARK)
    _c.alignment = _align(h="center", v="center")

    # S8:U8 – change columns, mid-green background
    ws.merge_cells("S8:U8")
    ws["S8"].fill = _fill(GREEN_MID)

    # ── Header Row 9: Individual column names ──────────────────────────────
    ws.row_dimensions[9].height = 40

    col_headers = {
        2:  "Benefit",
        3:  "Contract",
        4:  "Division",
        5:  "Class",
        6:  "Plan",
        7:  "Basis",
        8:  "",
        9:  "Lives",
        10: "Volumes/\nSalary",
        11: "Experience\nRate",
        12: "Deficit\nRecovery\nRate",
        13: "Total\nRate",
        14: "Total\nPremium ($)",
        15: "Experience\nRate",
        16: "Deficit\nRecovery\nRate",
        17: "Total\nRate",
        18: "Total\nPremium ($)",
        19: "Exp Rate\nChange",
        20: "DR\nPercentage",
        21: "Total Rate\nChange",
    }
    for col, lbl in col_headers.items():
        cell = ws.cell(row=9, column=col)
        cell.value = lbl
        if col in (11, 12, 13, 14):
            cell.fill = _fill(GREEN_HEADER)
            cell.font = _font(bold=True, size=8, color=WHITE)
        elif col in (15, 16, 17, 18):
            cell.fill = _fill(GREEN_DARK)
            cell.font = _font(bold=True, size=8, color=WHITE)
        elif col in (19, 20, 21):
            cell.fill = _fill(GREEN_MID)
            cell.font = _font(bold=True, size=8, color=WHITE)
        else:
            cell.fill = _fill(GREEN_MID)
            cell.font = _font(bold=True, size=8, color=WHITE)
        cell.alignment = _align(h="center", v="center", wrap=True)

    # ── Row number tracker (data starts at row 11) ─────────────────────────
    row = 11

    # ──────────────────────────────────────────────────────────────────────
    # BENEFIT SECTIONS
    # ──────────────────────────────────────────────────────────────────────

    def blank_row():
        nonlocal row
        row += 1

    def section(label, code, placeholder_rows=None):
        """Write one benefit section. Returns row cursor after writing."""
        nonlocal row

        section_rows = groups.get(code, [])
        if placeholder_rows and not section_rows:
            section_rows = placeholder_rows

        if not section_rows:
            return 0, 0

        _write_section_header(ws, row, label)
        row += 1

        curr_total_prem = 0
        prop_total_prem = 0
        total_lives     = 0
        total_volumes   = 0
        last_basis = ""

        for r in section_rows:
            r["benefit_code"] = code
            _write_data_row(ws, row, r)
            curr_total_prem += r.get("curr_premium", 0) or 0
            prop_total_prem += r.get("prop_premium", 0) or 0
            total_lives     += r.get("lives", 0) or 0
            total_volumes   += r.get("volumes", 0) or 0
            last_basis = r.get("basis_unit", "")
            row += 1

        blank_row()
        blank_row()

        # Compute blended total change
        total_chg = None
        if curr_total_prem:
            total_chg = round((prop_total_prem - curr_total_prem) / curr_total_prem, 6)

        # Split Single/Family lives for EHC/DENT sub-totals
        if code in ("EHC", "DENT"):
            singles = sum(r.get("lives", 0) for r in section_rows if r.get("basis_unit") == "Single")
            families = sum(r.get("lives", 0) for r in section_rows if r.get("basis_unit") == "Family")
            total_lives_display = singles + families
            ws.cell(row=row-1, column=7, value="Single")
            ws.cell(row=row-1, column=9, value=singles)
            ws.cell(row=row-2, column=7, value="Family")
            ws.cell(row=row-2, column=9, value=families)
        else:
            total_lives_display = total_lives

        _write_total_row(
            ws, row,
            f"Total {label}",
            total_lives_display,
            total_volumes or None,
            curr_total_prem,
            prop_total_prem,
            total_chg,
        )
        row += 1
        blank_row()

        return curr_total_prem, prop_total_prem

    # ── Placeholder row for zero-data benefits ────────────────────────────
    zero_life = lambda code, basis: [{
        "division": "0", "plans": "0", "lives": 0, "volumes": 0,
        "basis": "", "basis_unit": basis,
        "curr_exp_rate": 0, "curr_dr_rate": 0, "curr_total_rate": 0,
        "curr_premium": 0, "prop_exp_rate": 0, "prop_dr_rate": 0,
        "prop_total_rate": 0, "prop_premium": 0,
        "exp_rate_chg": 0, "dr_pct": 0, "total_rate_chg": None,
        "benefit_code": code,
    }]

    grand_curr = 0
    grand_prop = 0

    c, p = section("Basic Life",   "LIFE")
    grand_curr += c; grand_prop += p

    c, p = section("Dependent Life", "DEPL")
    grand_curr += c; grand_prop += p

    c, p = section("Optional Life", "OPTL")
    grand_curr += c; grand_prop += p

    c, p = section("Accidental Death & Dismemberment", "ADD")
    grand_curr += c; grand_prop += p

    c, p = section("Long Term Disability", "LTD")
    grand_curr += c; grand_prop += p

    c, p = section("Short Term Disability", "STD")
    grand_curr += c; grand_prop += p

    # EHC sub-sections
    def ehc_sub_section(label, rate_type_filter):
        nonlocal row
        _write_section_header(ws, row, label)
        row += 1
        ehc_rows = [r for r in groups.get("EHC", []) if r.get("basis_unit") == rate_type_filter]
        if not ehc_rows:
            ehc_rows = zero_life("EHC", rate_type_filter)
        c_total = 0; p_total = 0; lives_total = 0
        for r in ehc_rows:
            r["benefit_code"] = "EHC"
            _write_data_row(ws, row, r)
            c_total += r.get("curr_premium", 0)
            p_total += r.get("prop_premium", 0)
            lives_total += r.get("lives", 0)
            row += 1
        blank_row()
        blank_row()
        return c_total, p_total

    _write_section_header(ws, row, "Extended Health Care")
    row += 1
    # Write ALL EHC rows under one section header
    ehc_groups_rows = groups.get("EHC", [])
    if not ehc_groups_rows:
        ehc_groups_rows = zero_life("EHC", "Single") + zero_life("EHC", "Family")
    ehc_curr = 0; ehc_prop = 0; ehc_lives = 0
    for r in ehc_groups_rows:
        r["benefit_code"] = "EHC"
        _write_data_row(ws, row, r)
        ehc_curr += r.get("curr_premium", 0)
        ehc_prop += r.get("prop_premium", 0)
        ehc_lives += r.get("lives", 0)
        row += 1
    blank_row()
    blank_row()
    # Single / Family subtotals
    single_lives = sum(r.get("lives", 0) for r in ehc_groups_rows if r.get("basis_unit") == "Single")
    family_lives = sum(r.get("lives", 0) for r in ehc_groups_rows if r.get("basis_unit") == "Family")
    ws.cell(row=row, column=2, value=""); ws.cell(row=row, column=7, value="Single"); ws.cell(row=row, column=9, value=single_lives)
    row += 1
    ws.cell(row=row, column=2, value=""); ws.cell(row=row, column=7, value="Family"); ws.cell(row=row, column=9, value=family_lives)
    row += 1
    ehc_total_chg = round((ehc_prop - ehc_curr) / ehc_curr, 6) if ehc_curr else None
    _write_total_row(ws, row, "Total Extended Health Care", ehc_lives, None, ehc_curr, ehc_prop, ehc_total_chg)
    grand_curr += ehc_curr; grand_prop += ehc_prop
    row += 1; blank_row()

    # Hospital / Drugs / Vision / Pool Charges – only if data exists
    for lbl, code in [("Hospital", "HOSP"), ("Drugs", "DRUG"), ("Vision", "VIS"), ("Pool Charges", "POOL")]:
        c, p = section(lbl, code)
        grand_curr += c; grand_prop += p

    # Dental
    _write_section_header(ws, row, "Dental Care")
    row += 1
    dent_rows = groups.get("DENT", [])
    if not dent_rows:
        dent_rows = zero_life("DENT", "Single") + zero_life("DENT", "Family")
    dent_curr = 0; dent_prop = 0; dent_lives = 0
    for r in dent_rows:
        r["benefit_code"] = "DENT"
        _write_data_row(ws, row, r)
        dent_curr += r.get("curr_premium", 0)
        dent_prop += r.get("prop_premium", 0)
        dent_lives += r.get("lives", 0)
        row += 1
    blank_row()
    blank_row()
    single_lives = sum(r.get("lives", 0) for r in dent_rows if r.get("basis_unit") == "Single")
    family_lives = sum(r.get("lives", 0) for r in dent_rows if r.get("basis_unit") == "Family")
    ws.cell(row=row, column=7, value="Single"); ws.cell(row=row, column=9, value=single_lives); row += 1
    ws.cell(row=row, column=7, value="Family"); ws.cell(row=row, column=9, value=family_lives); row += 1
    dent_total_chg = round((dent_prop - dent_curr) / dent_curr, 6) if dent_curr else None
    _write_total_row(ws, row, "Total Dental Care", dent_lives, None, dent_curr, dent_prop, dent_total_chg)
    grand_curr += dent_curr; grand_prop += dent_prop
    row += 1; blank_row()

    # Critical Illness / Resilience – only if data exists
    for lbl, code in [("Critical Illness", "CI"), ("Resilience", "RES")]:
        c, p = section(lbl, code)
        grand_curr += c; grand_prop += p

    # ── Grand Total ────────────────────────────────────────────────────────
    blank_row()
    gt_chg = round((grand_prop - grand_curr) / grand_curr, 6) if grand_curr else None
    _write_section_header(ws, row, "GRAND TOTAL")
    ws.cell(row=row, column=14, value=round(grand_curr, 2))
    ws.cell(row=row, column=18, value=round(grand_prop, 2))
    ws.cell(row=row, column=19, value=gt_chg)
    ws.cell(row=row, column=21, value=gt_chg)
    for c in range(2, 22):
        ws.cell(row=row, column=c).font  = _font(bold=True, size=10, color=WHITE)
        ws.cell(row=row, column=c).fill  = _fill(GREEN_DARK)
    row += 1

    # ── Freeze panes and finalize ──────────────────────────────────────────
    ws.freeze_panes = "B10"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_sold_rate_sheet(report_data: dict, sold_rows: list) -> bytes:
    """
    Generate the Sold Rate Sheet Excel (ManuConnect format).
    One row per Plan × Benefit × Coverage Type with current & proposed rates.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sold Rate Sheet"
    ws.sheet_view.showGridLines = False

    client_name     = report_data.get("client_name", "")
    contract_numbers = report_data.get("contract_numbers", [])
    contract_no     = str(contract_numbers[0]) if contract_numbers else ""
    period_str      = report_data.get("period_str", "")

    # Parse effective date from period string → "01Feb2026" format
    import re
    eff_date = "01Feb2026"
    m = re.search(r"(\d{4})", period_str)
    if m:
        eff_date = f"01Feb{m.group(1)}"

    # ── Column definitions ────────────────────────────────────────────────
    HEADERS = [
        "System of Origin", "Group Name", "Group Number",
        "Aggregated Description", "Account", "Class", "Status",
        "Benefit Name", "Benefit Category", "Coverage Type",
        "RGO Proposed Rate", "Current Rate", "RGO Proposed Rate\nEffective Date",
    ]
    COL_WIDTHS = [16, 20, 14, 22, 9, 8, 12, 30, 22, 14, 16, 14, 20]

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Header row (yellow, bold, with auto-filter) ───────────────────────
    YELLOW = "FFD700"
    ws.row_dimensions[1].height = 30
    for col, hdr in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font      = Font(name="Calibri", bold=True, size=10, color="000000")
        cell.fill      = PatternFill("solid", fgColor=YELLOW)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = Border(
            bottom=Side(style="thin", color="999999"),
            right=Side(style="thin", color="CCCCCC"),
        )
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    # ── Data rows ─────────────────────────────────────────────────────────
    for r_idx, r in enumerate(sold_rows, 2):
        shade = "F9F9F9" if r_idx % 2 == 0 else "FFFFFF"
        vals = [
            "ManuConnect",
            client_name,
            contract_no,
            "",
            "000",
            r["plan"],
            "Confirmed",
            r["benefit_name"],
            r["benefit_category"],
            r["coverage_type"],
            r["proposed_rate"],
            r["current_rate"],
            eff_date,
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=shade)
            cell.alignment = Alignment(
                horizontal="right" if col in (11, 12) else "left",
                vertical="center"
            )
            if col in (11, 12):
                cell.number_format = "0.000"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
